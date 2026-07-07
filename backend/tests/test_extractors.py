"""
Tests for PDF and Excel extractors, subject detection utilities, and Excel generator.
All fixtures are built in-memory — no external files required.
"""
import io
from unittest.mock import patch

import openpyxl
import pytest

from app.schemas.schemas import ExtractedDataSchema, StudentRecord, StyleConfig, SubjectEntry


# ── Minimal PDF builder ──────────────────────────────────────────────────────
# Constructs a 1-page PDF containing the given ASCII text without external deps.

def _build_pdf(text: str) -> bytes:
    """Return bytes of a valid PDF containing *text* on one page."""
    stream = f"BT /F1 10 Tf 40 750 Td ({text}) Tj ET".encode("latin-1")
    n = len(stream)

    obj1 = b"1 0 obj\n<< /Type /Catalog /Pages 2 0 R >>\nendobj\n"
    obj2 = b"2 0 obj\n<< /Type /Pages /Kids [3 0 R] /Count 1 >>\nendobj\n"
    obj3 = (
        b"3 0 obj\n<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792]"
        b" /Contents 4 0 R"
        b" /Resources << /Font << /F1 << /Type /Font /Subtype /Type1"
        b" /BaseFont /Courier >> >> >> >>\nendobj\n"
    )
    obj4 = f"4 0 obj\n<< /Length {n} >>\nstream\n".encode() + stream + b"\nendstream\nendobj\n"

    header = b"%PDF-1.4\n"
    body = obj1 + obj2 + obj3 + obj4

    # Compute xref byte offsets
    o1 = len(header)
    o2 = o1 + len(obj1)
    o3 = o2 + len(obj2)
    o4 = o3 + len(obj3)
    xref_pos = len(header) + len(body)

    xref = (
        f"xref\n0 5\n"
        f"0000000000 65535 f \n"
        f"{o1:010d} 00000 n \n"
        f"{o2:010d} 00000 n \n"
        f"{o3:010d} 00000 n \n"
        f"{o4:010d} 00000 n \n"
        f"trailer\n<< /Size 5 /Root 1 0 R >>\n"
        f"startxref\n{xref_pos}\n%%EOF"
    ).encode()

    return header + body + xref


# ── Attestation sheet text ────────────────────────────────────────────────────

_ATTEST_TEXT = (
    "UNIVERSITY ATTESTATION SHEET  B.Com 3rd Semester  Exam: Nov 2024\n"
    "Roll No: 10001  MBAN301 Business Mathematics  MBAN302 Accountancy\n"
    "Roll No: 10002  MBAN301 Business Mathematics  MBAN303 Economics\n"
    "Roll No: 10003  MBAN302 Accountancy  MBAN303 Economics\n"
    "Roll No: 10004  MBAN301 Business Mathematics\n"
)


# ── PDF extractor tests ───────────────────────────────────────────────────────

def test_pdf_extractor_attestation_sheet():
    """Patch page-text extraction to return one student per page (real attestation format)."""
    from app.services.extractors.pdf_extractor import extract_from_pdf

    # Real attestation sheets have one student per page — simulate that here
    mock_pages = [
        "Roll No: 10001  MBAN301 Business Mathematics  MBAN302 Accountancy",
        "Roll No: 10002  MBAN301 Business Mathematics  MBAN303 Economics",
        "Roll No: 10003  MBAN302 Accountancy  MBAN303 Economics",
    ]

    with patch("app.services.extractors.pdf_extractor._extract_page_texts", return_value=mock_pages):
        students, subjects, text_sample = extract_from_pdf(b"dummy", "attest.pdf")

    assert len(students) == 3, f"Expected 3 students, got {len(students)}"
    roll_numbers = {s.roll_number for s in students}
    assert "10001" in roll_numbers
    assert "10002" in roll_numbers
    assert "10003" in roll_numbers

    assert len(subjects) > 0, "Expected at least one subject detected"
    subject_codes = set(subjects.keys())
    assert "MBAN301" in subject_codes or "MBAN302" in subject_codes

    assert text_sample  # non-empty sample


def test_pdf_extractor_empty_pdf():
    """An empty page list returns empty students/subjects without raising."""
    from app.services.extractors.pdf_extractor import extract_from_pdf

    with patch("app.services.extractors.pdf_extractor._extract_page_texts", return_value=[]):
        students, subjects, text_sample = extract_from_pdf(b"dummy", "empty.pdf")

    assert students == []
    assert subjects == {}
    assert text_sample == ""


def test_pdf_extractor_no_roll_numbers():
    """Pages with no roll-number patterns yield no students but may yield subjects."""
    from app.services.extractors.pdf_extractor import extract_from_pdf

    pages = ["Department of Commerce\nMBAN301 Business Mathematics\nMBAN302 Accountancy"]
    with patch("app.services.extractors.pdf_extractor._extract_page_texts", return_value=pages):
        students, subjects, _ = extract_from_pdf(b"dummy", "subjects_only.pdf")

    assert students == []
    assert "MBAN301" in subjects or "MBAN302" in subjects


# ── Excel extractor tests ─────────────────────────────────────────────────────

def _make_matrix_excel() -> bytes:
    """Format A: header row has subject codes as column headers."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Header row: roll-no column + subject code columns
    ws.append(["Roll No", "BCOM301", "BCOM302", "BCOM303"])
    # Student rows: 1 = enrolled, 0/None = not enrolled
    ws.append(["2001", 1, 1, 0])
    ws.append(["2002", 1, 0, 1])
    ws.append(["2003", 0, 1, 1])
    ws.append(["2004", 1, 1, 1])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_flat_excel() -> bytes:
    """Format B: one column has comma/space-separated paper codes per student."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws.append(["S.No", "Roll No", "Name", "Paper Codes"])
    ws.append([1, "3001", "Alice", "PHY401 PHY402 PHY403"])
    ws.append([2, "3002", "Bob",   "PHY401 PHY403"])
    ws.append([3, "3003", "Carol", "PHY402 PHY403 PHY404"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_excel_extractor_format_a():
    """Matrix format: roll numbers map to columns whose header is a subject code."""
    from app.services.extractors.excel_extractor import extract_from_excel

    xlsx = _make_matrix_excel()
    students, subjects, text_sample = extract_from_excel(xlsx, "matrix.xlsx")

    assert len(students) == 4
    rolls = {s.roll_number for s in students}
    assert "2001" in rolls and "2004" in rolls

    # Student 2001 enrolled in BCOM301 and BCOM302 (truthy), not BCOM303 (0)
    s2001 = next(s for s in students if s.roll_number == "2001")
    assert "BCOM301" in s2001.subjects
    assert "BCOM302" in s2001.subjects
    assert "BCOM303" not in s2001.subjects

    # All three codes should be in subject map
    assert "BCOM301" in subjects
    assert "BCOM302" in subjects
    assert "BCOM303" in subjects

    assert text_sample  # non-empty


def test_excel_extractor_format_b():
    """Flat-list format: one column holds space-separated subject codes."""
    from app.services.extractors.excel_extractor import extract_from_excel

    xlsx = _make_flat_excel()
    students, subjects, text_sample = extract_from_excel(xlsx, "flat.xlsx")

    assert len(students) == 3
    rolls = {s.roll_number for s in students}
    assert "3001" in rolls and "3003" in rolls

    s3001 = next(s for s in students if s.roll_number == "3001")
    assert "PHY401" in s3001.subjects
    assert "PHY402" in s3001.subjects
    assert "PHY403" in s3001.subjects

    s3002 = next(s for s in students if s.roll_number == "3002")
    assert "PHY401" in s3002.subjects
    assert "PHY403" in s3002.subjects
    assert "PHY402" not in s3002.subjects

    assert text_sample


def test_excel_extractor_empty_workbook():
    """An empty workbook returns empty results without raising."""
    from app.services.extractors.excel_extractor import extract_from_excel

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    buf = io.BytesIO()
    wb.save(buf)

    students, subjects, text_sample = extract_from_excel(buf.getvalue(), "empty.xlsx")
    assert students == []
    assert subjects == {}


def test_excel_extractor_corrupted_bytes():
    """Corrupted bytes raise RuntimeError (not an unhandled crash)."""
    from app.services.extractors.excel_extractor import extract_from_excel

    with pytest.raises(RuntimeError, match="Cannot open Excel file"):
        extract_from_excel(b"not an xlsx file at all", "bad.xlsx")


# ── Subject detection tests ───────────────────────────────────────────────────

def test_subject_detection_standard_codes():
    from app.utils.subject_utils import extract_all_subjects

    text = "Student enrolled in MBAN301 Business Mathematics and MBAN302 Accountancy"
    subjects = extract_all_subjects(text)

    assert "MBAN301" in subjects
    assert "MBAN302" in subjects


def test_subject_detection_standalone_codes():
    """Subject codes that appear without a following name are still detected."""
    from app.utils.subject_utils import extract_all_subjects

    text = "Papers: CS401, CS402, CS403 are registered"
    subjects = extract_all_subjects(text)

    assert "CS401" in subjects
    assert "CS402" in subjects
    assert "CS403" in subjects


def test_subject_detection_no_codes():
    """Text with no subject code patterns returns an empty dict."""
    from app.utils.subject_utils import extract_all_subjects

    subjects = extract_all_subjects("This is a plain text with no subject codes.")
    assert subjects == {}


def test_subject_detection_mixed():
    """Named (CODE - Name) and unnamed codes coexist correctly."""
    from app.utils.subject_utils import extract_all_subjects

    # _PAIR_RE requires "CODE - Name" or "CODE: Name" format for named extraction
    text = "ENGG201 - Engineering Drawing\nENGG202\nBCA101 - Computer Fundamentals"
    subjects = extract_all_subjects(text)

    assert "ENGG201" in subjects
    assert "ENGG202" in subjects
    assert "BCA101" in subjects
    # Named ones get non-empty names via the paired pattern
    assert subjects.get("ENGG201")  # "Engineering Drawing"
    assert subjects.get("BCA101")   # "Computer Fundamentals"
    # Unnamed code gets empty string
    assert subjects.get("ENGG202") == ""


def test_subject_detection_numeric_codes():
    from app.utils.subject_utils import extract_all_subjects

    text = (
        "210236 - Organisational Behaviour : Mi\n"
        "210242 - Financial Management : Mj\n"
        "Address: Jabalpur PIN-482001\n"
        "Address 2: pin 482002\n"
        "Address 3: PIN : 482003\n"
        "Address 4: pin - 482004\n"
        "Address 5: pin: 482005\n"
    )
    subjects = extract_all_subjects(text)

    assert "210236" in subjects
    assert "210242" in subjects
    assert subjects["210236"] == "Organisational Behaviour Mi"
    assert subjects["210242"] == "Financial Management Mj"

    for p in ["482001", "482002", "482003", "482004", "482005"]:
        assert p not in subjects


# ── Excel generator tests ─────────────────────────────────────────────────────

def _mock_extracted_data() -> ExtractedDataSchema:
    subjects = [
        SubjectEntry(code="MBAN301", name="Business Mathematics"),
        SubjectEntry(code="MBAN302", name="Accountancy"),
        SubjectEntry(code="MBAN303", name="Economics"),
    ]
    students = [
        StudentRecord(roll_number="10001", subjects=["MBAN301", "MBAN302"]),
        StudentRecord(roll_number="10002", subjects=["MBAN301", "MBAN303"]),
        StudentRecord(roll_number="10003", subjects=["MBAN302", "MBAN303"]),
        StudentRecord(roll_number="10004", subjects=["MBAN301", "MBAN302", "MBAN303"]),
        StudentRecord(roll_number="10005", subjects=["MBAN301"]),
    ]
    return ExtractedDataSchema(
        students=students,
        subjects=subjects,
        source_file="test.pdf",
        total_students=len(students),
        document_type="attestation_sheet",
        course="B.Com",
        semester="3rd Semester",
        exam_name="Nov 2024",
        ai_confidence=0.92,
    )


def test_excel_generator_returns_bytes():
    from app.services.generators.excel_generator import generate_excel

    data = _mock_extracted_data()
    result = generate_excel(data, StyleConfig(), "test-output")
    assert isinstance(result, bytes)
    assert len(result) > 0


def test_excel_generator_valid_workbook():
    """Generated bytes must be openable by openpyxl."""
    from app.services.generators.excel_generator import generate_excel

    data = _mock_extracted_data()
    xlsx_bytes = generate_excel(data, StyleConfig(), "test-output")

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    assert wb is not None


def test_excel_generator_sheet_names():
    """Output workbook must have both expected sheets."""
    from app.services.generators.excel_generator import generate_excel

    data = _mock_extracted_data()
    xlsx_bytes = generate_excel(data, StyleConfig(), "test-output")
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))

    assert "Subject-wise Roll Number List" in wb.sheetnames
    assert "Summary" in wb.sheetnames


def test_excel_generator_row_counts():
    """Sheet 1 must have header rows + data rows + count row; student data must appear."""
    from app.services.generators.excel_generator import generate_excel

    data = _mock_extracted_data()
    xlsx_bytes = generate_excel(data, StyleConfig(), "test-output")
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["Subject-wise Roll Number List"]

    # Row 1 = merged title, Row 2 = subject headers
    # Data rows start at row 3; max rolls in any subject = 4 (MBAN301)
    # Count row = row 3 + 4 = row 7
    title_cell = ws.cell(row=1, column=1).value
    assert title_cell and "ROLL NUMBER" in title_cell.upper()

    # Count how many subject columns exist (columns B onward)
    n_subjects = len(data.subjects)
    assert ws.cell(row=2, column=2).value is not None  # first subject header
    assert ws.cell(row=2, column=n_subjects + 1).value is not None

    # Verify data cells contain roll numbers from our students
    all_values = set()
    for row in ws.iter_rows(min_row=3, values_only=True):
        for cell in row:
            if cell is not None and str(cell).startswith("100"):
                all_values.add(str(cell))

    assert "10001" in all_values
    assert "10004" in all_values


def test_excel_generator_summary_sheet():
    """Summary sheet must list all subjects with student counts."""
    from app.services.generators.excel_generator import generate_excel

    data = _mock_extracted_data()
    xlsx_bytes = generate_excel(data, StyleConfig(), "test-output")
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["Summary"]

    # Collect all cell values as strings for easy searching
    all_values = [
        str(ws.cell(row=r, column=c).value or "")
        for r in range(1, ws.max_row + 1)
        for c in range(1, ws.max_column + 1)
    ]
    joined = " ".join(all_values)

    assert "MBAN301" in joined
    assert "MBAN302" in joined
    assert "MBAN303" in joined
    assert "B.Com" in joined


def test_excel_generator_custom_style():
    """Custom header color is applied without errors."""
    from app.services.generators.excel_generator import generate_excel

    data = _mock_extracted_data()
    style = StyleConfig(header_bg_color="#2C3E50", font_size=11, column_width=30)
    xlsx_bytes = generate_excel(data, style, "styled-output")

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["Subject-wise Roll Number List"]
    fill = ws.cell(row=1, column=1).fill
    # openpyxl stores fgColor as ARGB (FF prefix + hex)
    assert fill.fgColor.rgb.upper().endswith("2C3E50")


def test_excel_generator_no_students():
    """Generator must not crash when there are no students."""
    from app.services.generators.excel_generator import generate_excel

    data = ExtractedDataSchema(
        students=[],
        subjects=[SubjectEntry(code="MBAN301", name="Business Mathematics")],
        source_file="empty.pdf",
        total_students=0,
        document_type="unknown",
        ai_confidence=0.0,
    )
    xlsx_bytes = generate_excel(data, StyleConfig(), "empty-output")
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    assert "Subject-wise Roll Number List" in wb.sheetnames
