"""
Tests for multi-file aggregation (merge + dedupe) and output-time roll number
sorting. All pure logic — no files, DB, or network needed.
"""
import io

import openpyxl

from app.schemas.schemas import ExtractedDataSchema, StudentRecord, StyleConfig, SubjectEntry


# ── sort_roll_numbers ─────────────────────────────────────────────────────────

def test_sort_rolls_all_numeric_sorts_by_value():
    from app.utils.subject_utils import sort_roll_numbers

    assert sort_roll_numbers(["100", "23", "5", "1000"]) == ["5", "23", "100", "1000"]


def test_sort_rolls_natural_alphanumeric():
    """Embedded numbers must order by value: ...4 before ...20 before ...100."""
    from app.utils.subject_utils import sort_roll_numbers

    rolls = ["R22CS100", "R22CS4", "R22CS20"]
    assert sort_roll_numbers(rolls) == ["R22CS4", "R22CS20", "R22CS100"]


def test_sort_rolls_mixed_prefixes():
    from app.utils.subject_utils import sort_roll_numbers

    rolls = ["B2", "A10", "A2", "B1"]
    assert sort_roll_numbers(rolls) == ["A2", "A10", "B1", "B2"]


def test_sort_rolls_empty_and_single():
    from app.utils.subject_utils import sort_roll_numbers

    assert sort_roll_numbers([]) == []
    assert sort_roll_numbers(["42"]) == ["42"]


def test_sort_rolls_returns_new_list_no_mutation():
    from app.utils.subject_utils import sort_roll_numbers

    original = ["300", "100", "200"]
    result = sort_roll_numbers(original)
    assert result == ["100", "200", "300"]
    assert original == ["300", "100", "200"]  # input untouched
    assert result is not original


def test_sort_rolls_stable_for_equal_keys():
    """'007' and '7' compare equal as ints — original order must be kept."""
    from app.utils.subject_utils import sort_roll_numbers

    assert sort_roll_numbers(["007", "7", "3"]) == ["3", "007", "7"]


# ── build_subject_roll_map sorts per subject ─────────────────────────────────

def test_build_subject_roll_map_sorted_ascending():
    from app.utils.subject_utils import build_subject_roll_map

    students = [
        StudentRecord(roll_number="10030", subjects=["CS401"]),
        StudentRecord(roll_number="10002", subjects=["CS401", "CS402"]),
        StudentRecord(roll_number="10100", subjects=["CS402"]),
        StudentRecord(roll_number="9999", subjects=["CS401"]),
    ]
    subjects = [SubjectEntry(code="CS401", name=""), SubjectEntry(code="CS402", name="")]

    roll_map = build_subject_roll_map(students, subjects)

    assert roll_map["CS401"] == ["9999", "10002", "10030"]  # numeric, not lexicographic
    assert roll_map["CS402"] == ["10002", "10100"]
    # Stored student order untouched
    assert [s.roll_number for s in students] == ["10030", "10002", "10100", "9999"]


# ── merge_subject_maps ────────────────────────────────────────────────────────

def test_merge_subject_maps_later_files_fill_missing_names():
    from app.services.pipeline.processor import merge_subject_maps

    merged, warnings = merge_subject_maps([
        {"CS401": "", "CS402": "Databases"},
        {"CS401": "Operating Systems", "CS403": ""},
    ])
    assert merged == {"CS401": "Operating Systems", "CS402": "Databases", "CS403": ""}
    assert warnings == []


def test_merge_subject_maps_conflict_keeps_longer_name_and_warns():
    from app.services.pipeline.processor import merge_subject_maps

    merged, warnings = merge_subject_maps([
        {"CS401": "OS"},
        {"CS401": "Operating Systems"},
    ])
    assert merged["CS401"] == "Operating Systems"
    assert len(warnings) == 1
    assert "CS401" in warnings[0]


# ── aggregate_students ────────────────────────────────────────────────────────

def test_aggregate_dedupes_roll_subject_pairs_across_files():
    from app.services.pipeline.processor import aggregate_students

    file1 = [
        StudentRecord(roll_number="1001", subjects=["CS401", "CS402"]),
        StudentRecord(roll_number="1002", subjects=["CS401"]),
    ]
    file2 = [
        StudentRecord(roll_number="1001", subjects=["CS401", "CS403"]),  # CS401 dup
        StudentRecord(roll_number="1003", subjects=["CS402"]),
    ]

    students, duplicates = aggregate_students([file1, file2], dedupe=True)

    assert duplicates == 1  # only (1001, CS401) repeated
    by_roll = {s.roll_number: s.subjects for s in students}
    assert len(students) == 3
    assert by_roll["1001"] == ["CS401", "CS402", "CS403"]  # merged + sorted
    assert by_roll["1002"] == ["CS401"]
    assert by_roll["1003"] == ["CS402"]


def test_aggregate_identical_files_full_overlap():
    from app.services.pipeline.processor import aggregate_students

    file1 = [StudentRecord(roll_number="1001", subjects=["CS401", "CS402"])]
    students, duplicates = aggregate_students([file1, list(file1)], dedupe=True)

    assert len(students) == 1
    assert duplicates == 2  # both pairs seen twice


def test_aggregate_single_file_passthrough():
    """The n=1 case must come out unchanged (extractors already emit unique
    rolls with sorted subjects)."""
    from app.services.pipeline.processor import aggregate_students

    file1 = [
        StudentRecord(roll_number="1001", subjects=["CS401", "CS402"]),
        StudentRecord(roll_number="1002", subjects=["CS403"]),
    ]
    students, duplicates = aggregate_students([file1], dedupe=True)

    assert duplicates == 0
    assert [s.model_dump() for s in students] == [s.model_dump() for s in file1]


def test_aggregate_dedupe_off_keeps_everything():
    from app.services.pipeline.processor import aggregate_students

    file1 = [StudentRecord(roll_number="1001", subjects=["CS401"])]
    file2 = [StudentRecord(roll_number="1001", subjects=["CS401"])]
    students, duplicates = aggregate_students([file1, file2], dedupe=False)

    assert len(students) == 2
    assert duplicates == 0


# ── Excel generator receives (and preserves) sorted columns ──────────────────

def test_generator_columns_are_ascending():
    from app.services.generators.excel_generator import generate_excel

    students = [
        StudentRecord(roll_number="10100", subjects=["CS401"]),
        StudentRecord(roll_number="10002", subjects=["CS401"]),
        StudentRecord(roll_number="9999", subjects=["CS401"]),
        StudentRecord(roll_number="10030", subjects=["CS401"]),
    ]
    data = ExtractedDataSchema(
        students=students,
        subjects=[SubjectEntry(code="CS401", name="Operating Systems")],
        source_file="test.pdf",
        total_students=len(students),
        document_type="attestation_sheet",
        ai_confidence=0.9,
    )
    xlsx_bytes = generate_excel(data, StyleConfig(), "sorted-output")
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["Subject-wise Roll Number List"]

    column_values = [
        str(ws.cell(row=r, column=2).value)
        for r in range(3, 3 + len(students))
        if ws.cell(row=r, column=2).value is not None
    ]
    assert column_values == ["9999", "10002", "10030", "10100"]
