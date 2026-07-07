import re
import logging
from io import BytesIO

import openpyxl

from app.schemas.schemas import StudentRecord
from app.utils.subject_utils import extract_all_subjects, normalize_subject_name

logger = logging.getLogger(__name__)

_CODE_RE = re.compile(r"\b([A-Z]{2,6}\d{3,4}|\d{5,6})\b")

# Column header names that indicate the roll number column
_ROLL_HEADER_NAMES = frozenset(
    {"roll no", "rollno", "roll number", "roll_no", "enrollment", "enrolment", "roll"}
)


def extract_from_excel(
    file_bytes: bytes, filename: str
) -> tuple[list[StudentRecord], dict[str, str], str]:
    """Extract students, subjects, and a text sample from an Excel file.

    Auto-detects one of two layouts:
      Format A — Matrix: header row contains subject codes as columns, rows are students
      Format B — Flat list: a column holds comma/space-separated paper codes per student

    Returns:
        students    — list of StudentRecord
        subjects    — {code: name} dict
        text_sample — first ~3000 chars of row data as text, for AI classifier
    """
    students, subjects, text_sample, _row_count = extract_from_excel_with_stats(
        file_bytes, filename
    )
    return students, subjects, text_sample


def extract_from_excel_with_stats(
    file_bytes: bytes, filename: str
) -> tuple[list[StudentRecord], dict[str, str], str, int]:
    """Same extraction as extract_from_excel, plus the real data-row count for progress reporting."""
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise RuntimeError(f"Cannot open Excel file '{filename}': {exc}") from exc

    ws = _choose_sheet(wb)

    # Read all non-empty rows
    all_rows: list[tuple] = [
        row for row in ws.iter_rows(values_only=True)
        if any(cell is not None for cell in row)
    ]
    wb.close()

    if not all_rows:
        return [], {}, "", 0

    header_row = [str(c).strip() if c is not None else "" for c in all_rows[0]]
    data_rows = all_rows[1:]

    # Detect format by checking how many subject codes appear in the header
    header_codes = [h for h in header_row if _is_subject_code(h)]

    if len(header_codes) >= 2:
        students, subjects = _extract_format_a(header_row, data_rows, header_codes)
        logger.debug("Excel '%s' detected as Format A (matrix), %d codes in header", filename, len(header_codes))
    else:
        students, subjects = _extract_format_b(header_row, data_rows, filename)
        logger.debug("Excel '%s' detected as Format B (flat list)", filename)

    # Build text sample from the first few rows
    text_lines = ["\t".join(header_row)]
    for row in data_rows[:10]:
        text_lines.append("\t".join(str(c) if c is not None else "" for c in row))
    text_sample = "\n".join(text_lines)[:3000]

    return students, subjects, text_sample, len(data_rows)


# ── Format A: matrix / Split Subjects layout ─────────────────────────────────

def _extract_format_a(
    header_row: list[str],
    data_rows: list[tuple],
    header_codes: list[str],
) -> tuple[list[StudentRecord], dict[str, str]]:
    roll_col = _find_roll_col(header_row)

    # If no named roll column, pick the first column that holds numeric values
    if roll_col is None:
        for row in data_rows[:5]:
            for i, cell in enumerate(row):
                if cell is not None and str(cell).strip().isdigit():
                    roll_col = i
                    break
            if roll_col is not None:
                break
        roll_col = roll_col if roll_col is not None else 0

    # Map column index → subject code for every code column
    code_at: dict[int, str] = {
        idx: col
        for idx, col in enumerate(header_row)
        if _is_subject_code(col)
    }

    students: list[StudentRecord] = []
    for row in data_rows:
        if not row or all(c is None for c in row):
            continue

        roll = _cell_str(row, roll_col)
        if not roll or roll.lower() in {"none", "roll no", "roll number", "s.no", "sno"}:
            continue

        enrolled = [
            code
            for idx, code in code_at.items()
            if idx < len(row) and _truthy(row[idx])
        ]
        if enrolled:
            students.append(StudentRecord(roll_number=roll, subjects=sorted(enrolled)))

    # Subject map: codes are known from the header; names are empty (no name row present)
    subjects: dict[str, str] = {code: "" for code in header_codes}
    return students, subjects


# ── Format B: flat list layout ───────────────────────────────────────────────

def _extract_format_b(
    header_row: list[str],
    data_rows: list[tuple],
    filename: str,
) -> tuple[list[StudentRecord], dict[str, str]]:
    roll_col = _find_roll_col(header_row)
    if roll_col is None:
        # Heuristic: first column with numeric-looking values in top rows
        roll_col = _infer_roll_col(data_rows) or 1

    paper_col = _find_paper_col(header_row, data_rows)
    if paper_col is None:
        logger.warning("No paper code column found in '%s'; no students extracted", filename)
        return [], {}

    students: list[StudentRecord] = []
    all_subjects: dict[str, str] = {}

    for row in data_rows:
        if not row or all(c is None for c in row):
            continue

        roll = _cell_str(row, roll_col)
        if not roll or not roll.replace("-", "").replace("/", "").strip():
            continue
        # Skip header-like values repeated in data rows
        if roll.lower() in {"none", "roll no", "roll number", "s.no", "sno", "serial no"}:
            continue

        cell_text = _cell_str(row, paper_col) if paper_col < len(row) else ""
        codes = _CODE_RE.findall(cell_text)
        if not codes:
            continue

        students.append(StudentRecord(roll_number=roll, subjects=sorted(set(codes))))

        # Extract subject names from cell text if present
        cell_subjects = extract_all_subjects(cell_text)
        for code, name in cell_subjects.items():
            if name:
                all_subjects.setdefault(code, name)
            else:
                all_subjects.setdefault(code, "")

    return students, all_subjects


# ── Sheet / column helpers ────────────────────────────────────────────────────

def _choose_sheet(wb: openpyxl.Workbook):
    """Prefer 'Split Subjects' → 'Sheet1' → first sheet."""
    lower_to_orig = {n.lower(): n for n in wb.sheetnames}
    for preferred in ("split subjects", "sheet1"):
        if preferred in lower_to_orig:
            return wb[lower_to_orig[preferred]]
    return wb[wb.sheetnames[0]]


def _find_roll_col(header_row: list[str]) -> int | None:
    for i, h in enumerate(header_row):
        if h.lower().strip() in _ROLL_HEADER_NAMES:
            return i
    return None


def _infer_roll_col(data_rows: list[tuple]) -> int | None:
    """Return the index of the first column that holds numeric values in early rows."""
    for row in data_rows[:5]:
        for i, cell in enumerate(row):
            if cell is not None and str(cell).strip().isdigit():
                return i
    return None


def _find_paper_col(header_row: list[str], data_rows: list[tuple]) -> int | None:
    """Return the column index whose cells contain multiple subject codes."""
    for col_idx in range(len(header_row)):
        for row in data_rows[:15]:
            if col_idx >= len(row) or row[col_idx] is None:
                continue
            codes = _CODE_RE.findall(str(row[col_idx]))
            if len(codes) >= 2:
                return col_idx
    return None


def _is_subject_code(val: str) -> bool:
    return bool(re.fullmatch(r"[A-Z]{2,6}\d{3,4}|\d{5,6}", val.strip()))


def _truthy(val) -> bool:
    """Return True if val represents an enrolled / selected cell (1, 'yes', 'x', etc.)."""
    if val is None:
        return False
    s = str(val).strip().lower()
    return s not in {"", "0", "none", "no", "false", "-"}


def _cell_str(row: tuple, col: int) -> str:
    if col >= len(row) or row[col] is None:
        return ""
    return str(row[col]).strip()
