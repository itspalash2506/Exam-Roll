import io
from datetime import datetime
from typing import Dict, List

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.schemas.schemas import ExtractedDataSchema, StyleConfig
from app.utils.subject_utils import build_subject_roll_map

_BORDER_COLOR = "B0B0B0"


def generate_excel(
    extracted_data: ExtractedDataSchema,
    style_config: StyleConfig,
    output_filename: str,
) -> bytes:
    roll_map = _build_roll_map(extracted_data)
    subjects = extracted_data.subjects

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("Subject-wise Roll Number List")
    _build_sheet1(ws1, extracted_data, subjects, roll_map, style_config)

    ws2 = wb.create_sheet("Summary")
    _build_sheet2(ws2, extracted_data, subjects, roll_map, style_config)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Helpers ───────────────────────────────────────────────────────────────────

def _hex(color: str) -> str:
    return color.lstrip("#")


def _thin_border() -> Border:
    side = Side(style="thin", color=_BORDER_COLOR)
    return Border(left=side, right=side, top=side, bottom=side)


def _fill(color_hex: str) -> PatternFill:
    return PatternFill("solid", fgColor=color_hex)


def _build_roll_map(data: ExtractedDataSchema) -> Dict[str, List[str]]:
    # build_subject_roll_map returns each subject's rolls already sorted
    # ascending (numeric when purely numeric, natural sort otherwise) — the
    # writing logic below preserves that order and never re-sorts.
    return build_subject_roll_map(data.students, data.subjects)


# ── Sheet 1: Subject-wise Roll Number List ────────────────────────────────────

def _build_sheet1(
    ws,
    data: ExtractedDataSchema,
    subjects,
    roll_map: Dict[str, List[str]],
    style: StyleConfig,
) -> None:
    n_subjects = len(subjects)
    # Column A is the label column; subject columns start at B
    n_cols = n_subjects + 1
    border = _thin_border()

    hdr_bg = _hex(style.header_bg_color)
    hdr_fg = _hex(style.header_font_color)
    alt_bg = _hex(style.alt_row_color)
    cnt_bg = _hex(style.count_row_color)

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    subtitle_parts = [p for p in [data.exam_name, data.course, data.semester] if p]
    title_text = "SUBJECT-WISE ROLL NUMBER LIST"
    if subtitle_parts:
        title_text += "\n" + " | ".join(subtitle_parts)

    c = ws.cell(row=1, column=1, value=title_text)
    c.font = Font(name=style.font_name, bold=True, size=14, color="FFFFFF")
    c.fill = _fill(hdr_bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
    if n_cols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.row_dimensions[1].height = 40

    # ── Row 2: Column headers (one subject per column) ────────────────────────
    c = ws.cell(row=2, column=1, value="")
    c.fill = _fill(hdr_bg)
    c.font = Font(name=style.font_name, bold=True, size=style.font_size, color=hdr_fg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

    for i, subject in enumerate(subjects):
        col = i + 2
        c = ws.cell(row=2, column=col, value=f"{subject.code}\n{subject.name}")
        c.fill = _fill(hdr_bg)
        c.font = Font(name=style.font_name, bold=True, size=style.font_size, color=hdr_fg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[2].height = 50

    # ── Data rows (rows 3 … 3+data_rows-1) ───────────────────────────────────
    max_rolls = max((len(roll_map.get(s.code, [])) for s in subjects), default=0)
    # Keep at least 1 placeholder row so COUNTA has a valid non-backwards range
    data_rows = max(max_rolls, 1)

    for row_idx in range(data_rows):
        excel_row = row_idx + 3
        row_fill = _fill(alt_bg) if row_idx % 2 == 1 else _fill("FFFFFF")

        c = ws.cell(row=excel_row, column=1, value=None)
        c.fill = row_fill
        c.border = border

        for i, subject in enumerate(subjects):
            col = i + 2
            rolls = roll_map.get(subject.code, [])
            value = rolls[row_idx] if row_idx < len(rolls) else None
            c = ws.cell(row=excel_row, column=col, value=value)
            c.fill = row_fill
            c.font = Font(name=style.font_name, size=style.font_size)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
        ws.row_dimensions[excel_row].height = 18

    # ── Count row ─────────────────────────────────────────────────────────────
    count_row = data_rows + 3
    last_data_row = count_row - 1

    c = ws.cell(row=count_row, column=1, value="TOTAL")
    c.fill = _fill(cnt_bg)
    c.font = Font(name=style.font_name, bold=True, size=style.font_size)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

    for i, subject in enumerate(subjects):
        col = i + 2
        col_letter = get_column_letter(col)
        formula = f"=COUNTA({col_letter}3:{col_letter}{last_data_row})"
        c = ws.cell(row=count_row, column=col, value=formula)
        c.fill = _fill(cnt_bg)
        c.font = Font(name=style.font_name, bold=True, size=style.font_size)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws.row_dimensions[count_row].height = 22

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions[get_column_letter(1)].width = 10
    for i in range(n_subjects):
        ws.column_dimensions[get_column_letter(i + 2)].width = style.column_width

    # ── Freeze panes: title + header stay visible when scrolling ─────────────
    ws.freeze_panes = "A3"


# ── Sheet 2: Summary ──────────────────────────────────────────────────────────

def _build_sheet2(
    ws,
    data: ExtractedDataSchema,
    subjects,
    roll_map: Dict[str, List[str]],
    style: StyleConfig,
) -> None:
    border = _thin_border()
    hdr_bg = _hex(style.header_bg_color)
    hdr_fg = _hex(style.header_font_color)
    alt_bg = _hex(style.alt_row_color)
    cnt_bg = _hex(style.count_row_color)

    # ── Row 1: Sheet title ────────────────────────────────────────────────────
    c = ws.cell(row=1, column=1, value="EXAM ROLL SUMMARY")
    c.font = Font(name=style.font_name, bold=True, size=14, color="FFFFFF")
    c.fill = _fill(hdr_bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 35

    # ── Rows 2–7: Document metadata ───────────────────────────────────────────
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    metadata = [
        ("Document Type", (data.document_type or "").replace("_", " ").title()),
        ("Course", data.course or "—"),
        ("Semester", data.semester or "—"),
        ("Exam Name", data.exam_name or "—"),
        ("Total Students", str(data.total_students)),
        ("Generated On", now),
    ]
    for i, (label, value) in enumerate(metadata):
        row = i + 2
        lc = ws.cell(row=row, column=1, value=label + ":")
        lc.font = Font(name=style.font_name, bold=True, size=style.font_size)
        lc.alignment = Alignment(horizontal="right", vertical="center")
        vc = ws.cell(row=row, column=2, value=value)
        vc.font = Font(name=style.font_name, size=style.font_size)
        vc.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18

    # Row 8 is intentionally blank (separator before table)

    # ── Row 9: Table headers ──────────────────────────────────────────────────
    table_header_row = 9
    table_cols = ["S.No", "Subject Code", "Subject Name", "Students Enrolled"]
    for col, h in enumerate(table_cols, 1):
        c = ws.cell(row=table_header_row, column=col, value=h)
        c.fill = _fill(hdr_bg)
        c.font = Font(name=style.font_name, bold=True, size=style.font_size, color=hdr_fg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws.row_dimensions[table_header_row].height = 22

    # ── Data rows ─────────────────────────────────────────────────────────────
    sorted_subjects = sorted(subjects, key=lambda s: s.code)
    for idx, subject in enumerate(sorted_subjects):
        row = table_header_row + 1 + idx
        count = len(roll_map.get(subject.code, []))
        row_fill = _fill(alt_bg) if idx % 2 == 1 else None

        for col, val in enumerate([idx + 1, subject.code, subject.name, count], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(name=style.font_name, size=style.font_size)
            c.alignment = Alignment(
                horizontal="left" if col == 3 else "center",
                vertical="center",
            )
            c.border = border
            if row_fill:
                c.fill = row_fill
        ws.row_dimensions[row].height = 18

    # ── Total row ─────────────────────────────────────────────────────────────
    n_subjects = len(sorted_subjects)
    total_row = table_header_row + 1 + n_subjects
    data_start = table_header_row + 1
    data_end = total_row - 1

    sum_formula = f"=SUM(D{data_start}:D{data_end})" if n_subjects > 0 else 0
    for col, val in [(1, None), (2, "TOTAL"), (3, None), (4, sum_formula)]:
        c = ws.cell(row=total_row, column=col, value=val)
        c.fill = _fill(cnt_bg)
        c.font = Font(name=style.font_name, bold=True, size=style.font_size)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws.row_dimensions[total_row].height = 22

    # ── Note below table ──────────────────────────────────────────────────────
    note_row = total_row + 2
    note_text = (
        "* Students appearing in multiple subjects are counted once per subject. "
        "Total ≠ unique student count."
    )
    c = ws.cell(row=note_row, column=1, value=note_text)
    c.font = Font(
        name=style.font_name,
        size=max(style.font_size - 1, 8),
        italic=True,
        color="666666",
    )
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells(
        start_row=note_row, start_column=1, end_row=note_row, end_column=4
    )
    ws.row_dimensions[note_row].height = 30

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 20
